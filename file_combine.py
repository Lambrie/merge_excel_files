import openpyxl, argparse, os, pathlib, shutil, sys, logging
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import NamedStyle
from datetime import datetime

class AgentReport(object):
    def __init__(self, inAgent, inLoggedInTime, inSignedOnTime, inBreakTime, inIncomingCalls, inAnsweredIncoming,
                 inTalkTimeIncoming,inAbandonedIncomingCalls, inOutgoingCallsExternal):
        self.agent = inAgent
        self.loggedInTime = inLoggedInTime
        self.signedOnTime = inSignedOnTime
        self.breakTime = inBreakTime
        self.incomingCalls = inIncomingCalls
        self.answeredIncoming = inAnsweredIncoming
        self.talkTimeIncoming = inTalkTimeIncoming
        self.abandonedIncomingCalls = inAbandonedIncomingCalls
        self.outgoingCallsExternal = inOutgoingCallsExternal

    def getLineEntry(self, extractDate):
        return [self.agent,
                str(extractDate),
                self.loggedInTime,
                self.signedOnTime,
                self.breakTime,
                int(self.incomingCalls),
                int(self.answeredIncoming),
                self.talkTimeIncoming,
                int(self.abandonedIncomingCalls),
                int(self.outgoingCallsExternal)
                ]

    def getHeaders(self):
        return ["Agent", "Date", "Logged In Time [totTLogin / All]", "Signed On Time [totTSignon / All]",
         "Break Time [totTPause / All]", "Incoming Calls [totNNew<- / Tel]",
         "Answered Incoming [totNConv<- / Tel]", "Talk Time Incoming [totTConv<- / Tel]",
         "Abandoned Incoming Calls [totNAban<- / Tel]", "Outgoing Calls (External) [totNNew->Ext / Tel]"]

    def __str__(self):
        return "AgentReport"


class VolumeReport(object):
    def __init__(self, inTopic, inTotalIncomingCalls, inLostCalls, inNoAnswer, inAverageTalkTime, inLongestTalkTime,
                 inAverageSpeedToAnswer,inLongestAnswerTime,inAnswerRate,inTotalReroutedCalls):
        self.topic = inTopic
        self.totalIncomingCalls = inTotalIncomingCalls
        self.lostCalls = inLostCalls
        self.noAnswer = inNoAnswer
        self.averageTalkTime = inAverageTalkTime
        self.longestTalkTime = inLongestTalkTime
        self.averageSpeedToAnswer = inAverageSpeedToAnswer
        self.longestAnswerTime = inLongestAnswerTime
        self.answerRate = inAnswerRate
        self.totalReroutedCalls = inTotalReroutedCalls

    def getLineEntry(self, extractDate):
        return [self.topic,
             str(extractDate),
             int(self.totalIncomingCalls),
             int(self.lostCalls),
             int(self.noAnswer),
             str(self.averageTalkTime),
             str(self.longestTalkTime),
             str(self.averageSpeedToAnswer),
             str(self.longestAnswerTime),
             float(self.answerRate),
             int(self.totalReroutedCalls)
             ]

    def getHeaders(self):
        return ["Topic", "Date", "Total Incoming Calls [totNNew / Tel]", "Lost Calls [totNLost / Tel]",
                "No Answer (Timeout) [totNExp / Tel]", "Average Talk Time (ATT) [avgTConvAg / Tel]",
                "Longest Talk Time [maxTConvAg / Tel]", "Average Speed To Answer (ASA) [avgTConvWait / Tel]",
                "Longest Answer Time [maxTConvWait / Tel]", "Answer Rate (%) [%AcceptLevel / Tel]",
                "Total Rerouted Calls [totNRr / Tel]"]

    def __str__(self):
        return "VolumeReport"


class Report(object):
    def __init__(self, inFileName, inReportType): # ,inType, inCreated, inPeriod, inWeekly, inDaily, inResolution, inCycle
        self.fileName = inFileName
        self.reportType = None
        self.reportTypeOverride = inReportType
        self.dateCreated = None
        self.period = None
        self.weekly = None
        self.daily = None
        self.resolution = None
        self.cycle = None
        self.content = []

    def getHeaders(self):
        if len(self.content) > 0:
            return self.content[0].getHeaders()
        else:
            return []

    def extractDate(self):
        try:
            dt= datetime.strptime(self.period[:10],"%m/%d/%Y").strftime('%m/%d/%Y')
        except:
            logging.warning(f"Default Load Date -> {str(datetime.now().strftime('%m/%d/%Y'))}")
            return datetime.now().strftime('%m/%d/%Y')
        else:
            return dt

    def addReport(self, data):
        if self.reportTypeOverride == "Agent":
            report = AgentReport(data.get('agent',None),data.get('loggedIn',None),data.get('signedOn',None)
                                  ,data.get('breakTime',None),data.get('incomingCalls',None),
                                  data.get('answeredIncoming',None),data.get('talkTime',None),
                                  data.get('abondonedIncomingCalls',None),data.get('outgoingCalls',None))
            self.content.append(report)
        elif self.reportTypeOverride == "Volume":
            answerRate = data.get('answerRate', None)
            if "%" in answerRate:
                answerRate = answerRate.replace("%","")

            try:
                answerRate = float(answerRate)
            except TypeError:
                answerRate = None

            report = VolumeReport(data.get('topic', None), data.get('totalIncomingCalls', None),
                                      data.get('lostCalls', None)
                                      , data.get('noAnswer', None), data.get('averageTalkTime', None),
                                      data.get('longestTalkTime', None), data.get('averageSpeedToAnswer', None),
                                      data.get('longestAnswerTime', None), answerRate,
                                      data.get('totalReroutedCalls', None))
            self.content.append(report)


def loadWorkbook(file):
    if file.suffix in [".xlsx"]:
        try:
            xlsx = openpyxl.load_workbook(file.absolute())
        except FileNotFoundError as e:
            logging.error(f"{file.name} - {e}")
            return None
        except FileExistsError as e:
            logging.error(f"{file.name} - {e}")
            return None
        except PermissionError as e:
            logging.error(f"{file.name} - Please close open file")
            return None
        except Exception as e:
            logging.error(f"{file.name} FILE_ERROR - {e}")
            return None
        else:
            return xlsx
    else:
        logging.warning(f"{file.name} - Not a valid Excel file")
        return None


def getAllFileData(input_directory, reportType):
    '''
    :arg inpurt_directory
    :arg reportType

    Load all files from the input directory provided into a list of report objects based of the above
    defined Report class

    :returns len(inputFiles), reports

    This functions returns a count of all the .xlsx files recognised in the directory
    A list of all valid reports loaded with content
    '''

    print(f"File upload started from {input_directory}")
    logging.info(f"File upload started from {input_directory}")
    reports = []
    try:
        inputFiles = [os.path.join(input_directory, f) for f in os.listdir(input_directory) if "master" not in f.lower()]
    except FileNotFoundError as e:
        logging.error(f"{input_directory} - {e}")
        return 0, reports
    except FileExistsError as e:
        logging.error(f"{input_directory} - {e}")
        return 0, reports
    xlsxCount = 0
    for file in inputFiles:
        filePath = pathlib.Path(file)
        if filePath.suffix in [".xlsx"]:
            xlsxCount += 1
            xlsx = loadWorkbook(filePath)
            if xlsx is None: continue
            try:
                sheet = xlsx["Table"]
            except KeyError as e:
                logging.error(f"{e} - Ensure sheet 'Table' is present in file - {filePath.name}")
                continue
            report = Report(filePath.stem, reportType)
            for rowCount, row in enumerate(sheet.rows):
                content = {}
                if rowCount in range(0,9):
                    if "type" in str(row[0].value).lower():
                        report.reportType = str(row[1].value)
                    elif "created" in str(row[0].value).lower():
                        report.dateCreated = str(row[1].value)
                    elif "period" in str(row[0].value).lower():
                        report.period = str(row[1].value)
                    elif "weekly" in str(row[0].value).lower():
                        report.weekly = str(row[1].value)
                    elif "daily" in str(row[0].value).lower():
                        report.daily = str(row[1].value)
                    elif "resolution" in str(row[0].value).lower():
                        report.resolution = str(row[1].value)
                    elif "cycle" in str(row[0].value).lower():
                        report.cycle = str(row[1].value)
                    else: continue
                else:
                    if report.reportTypeOverride == "Volume":
                        content["topic"] = str(row[0].value)
                        content["totalIncomingCalls"] = str(row[1].value)
                        content["lostCalls"] = str(row[2].value)
                        content["noAnswer"] = str(row[3].value)
                        content["averageTalkTime"] = str(row[4].value)
                        content["longestTalkTime"] = str(row[5].value)
                        content["averageSpeedToAnswer"] = str(row[6].value)
                        content["longestAnswerTime"] = str(row[7].value)
                        try:
                            content["answerRate"] = str(row[8].value)
                        except IndexError:
                            content["answerRate"] = '0'
                        try:
                            content["totalReroutedCalls"] = str(row[9].value)
                        except IndexError:
                            content["totalReroutedCalls"] = '0'
                    elif report.reportTypeOverride == "Agent":
                        if "log." in str(row[0].value).lower():
                            break
                        content["agent"] = str(row[0].value)
                        content["loggedIn"] = str(row[1].value)
                        content["signedOn"] = str(row[2].value)
                        content["breakTime"] = str(row[3].value)
                        content["incomingCalls"] = str(row[4].value)
                        content["answeredIncoming"] = str(row[5].value)
                        content["talkTime"] = str(row[6].value)
                        content["abondonedIncomingCalls"] = str(row[7].value)
                        content["outgoingCalls"] = str(row[8].value)
                    report.addReport(content)
            reports.append(report)
            xlsx.close()

    return xlsxCount, reports


def mergeData(output_file, reports):
    '''
    :arg
    output_file - Receives argument specify location of Master.xlsx
    report - Accept a list of valid reports loaded from the getAllFileData function

    :returns A count of all successful created and failed records in the output_file

    '''
    logging.info(f"Merging results into {output_file}")
    filePath = pathlib.Path(output_file)
    try:
        xlsx = openpyxl.load_workbook(filePath.absolute())
    except FileNotFoundError as e:
        logging.warning(f"{filePath.name} - Not Found")
        xlsx = openpyxl.Workbook()
        logging.info(f"Created new workbook -> {filePath.name}")
    except PermissionError as e:
        logging.error(f"{filePath.name} - Please close open file")
        return 0,0
    except Exception as e:
        logging.error(f"{filePath.name} FILE_ERROR",exc_info=True)
        return 0,0

    try:
        tempReport = reports[0]
        date_style = NamedStyle(name='american_date_style', number_format='MM/DD/YYYY')
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        if 'Daily Agent Reports' not in xlsx.sheetnames:
            xlsx.create_sheet('Daily Agent Reports')
            sheet = xlsx["Daily Agent Reports"]
            table = Table(displayName="AgentResults", ref="A1:J1")
            table.tableStyleInfo = style
            sheet.add_table(table)
            sheet.append(tempReport.getHeaders())
        elif 'Daily Volume Reports' not in xlsx.sheetnames:
            xlsx.create_sheet('Daily Volume Reports')
            sheet = xlsx["Daily Volume Reports"]
            table = Table(displayName="VolumeResults", ref="A1:K1")
            table.tableStyleInfo = style
            sheet.add_table(table)
            sheet.append(tempReport.getHeaders())
        else:
            if tempReport.reportTypeOverride == "Agent":
                sheet = xlsx["Daily Agent Reports"]
                table = sheet.tables["AgentResults"]
            else:
                sheet = xlsx["Daily Volume Reports"]
                table = sheet.tables["VolumeResults"]


    except IndexError:
        logging.error(f"No headers data available for {filePath.name}")
        return 0, 0

    successCounter = 0
    failedCounter = 0
    lastrow = table.ref.split(":")[-1]
    lastrowColumn = lastrow[0]
    lastrowIndex = int(lastrow[1:])
    for report in reports:
        for data in report.content:
            try:
                sheet.append(data.getLineEntry(report.extractDate()))
                successCounter +=1
            except Exception as e:
                logging.error("Exception occurred", exc_info=True)
                failedCounter +=1
    table.ref = f"A1:{lastrowColumn}{lastrowIndex + successCounter}"

    # Format Date Cells
    # for row in sheet[2:sheet.max_row]:  # skip the header
    #     cell = row[1]  # column B
    #     cell.number_format = 'dd/mm/yy'
    ##

    xlsx.save(filename=filePath)
    xlsx.close()
    return successCounter, failedCounter


def removeInputFile(input_directory,archive_directory=None, archive=False):
    '''
    :arg
    inpurt_directory - specifies the input directory from where files were consumed

    This function will delete all files in the input_directory after the Master.xlsx was updated, if archive is set to False
    If archive flag is set to true an archive directory will be expected, and all input files will only be moved to
    an archive location.
    '''
    if archive:
        if os.path.isdir(archive_directory) == False:
            logging.error(f"Invalid archive directory {archive_directory}")
            pass

        for filename in os.listdir(input_directory):
            if "master" in filename.lower(): continue
            file_path = os.path.join(input_directory, filename)
            try:
                if os.path.isfile(file_path) and os.path.isdir(archive_directory):
                    shutil.move(file_path, archive_directory)
                    logging.info(f"{filename} archived to {archive_directory}")
            except shutil.Error:
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.unlink(file_path)
                logging.warning(f'{file_path} already exist in archive, therfore {file_path} will be deleted and not archived', exc_info=True)
            except Exception as e:
                logging.error(f'Failed to archive file {file_path}', exc_info=True)
    else:
        for filename in os.listdir(input_directory):
            if "master" in filename.lower(): continue
            file_path = os.path.join(input_directory, filename)
            try:
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.unlink(file_path)
                elif os.path.isdir(file_path):
                    shutil.rmtree(file_path)
            except Exception as e:
                logging.error(f'Failed to delete file {file_path}', exc_info=True)
            else:
                logging.info(f"{filename} removed from directory - {file_path}")


def performFileMerge(files):
    logging.info(f"Load {files.get('type')} files")
    inputFileCount, reports = getAllFileData(files.get("input"), files.get("type"))
    logging.info(f"{str(len(reports))}/{str(inputFileCount)} files loaded")
    if inputFileCount > 0:
        success, failed = mergeData(files.get("output"),
                                    [report for report in reports if report.reportTypeOverride == files.get("type")])
        if success > 0:
            logging.info(f"{str(success)} rows merged successfully and {failed} records failed to merge")
            removeInputFile(files.get("input"), archive_directory=files.get("archive"),archive=True)


def run(argv=None):
    parser = argparse.ArgumentParser()
    parser.add_argument(
        '--agent_input_directory',
        dest='agent_input_directory',
        required=True,
        help='Directory of weekly input files for Agent Reports')
    parser.add_argument(
        '--agent_archive_directory',
        dest='agent_archive_directory',
        required=True,
        help='Archive directory of weekly input files for Agent Reports')
    parser.add_argument(
        '--volume_input_directory',
        dest='volume_input_directory',
        required=True,
        help='Directory of weekly input files for Volume Reports')
    parser.add_argument(
        '--volume_archive_directory',
        dest='volume_archive_directory',
        required=True,
        help='Archive directory of weekly input files for Volume Reports')
    parser.add_argument(
        '--master_output_file',
        dest='master_output_file',
        required=True,
        help='Output file name and location to write results to')
    known_args, _ = parser.parse_known_args(argv)

    logging.basicConfig(filename='file_combine.log', filemode='w', level=logging.INFO,
                        format='%(asctime)s - %(levelname)s - %(message)s',
                        datefmt='%d-%b-%y %H:%M:%S')

    # Check for valid arguments
    if os.path.isdir(known_args.agent_input_directory) == False:
        logging.error(f"Invalid directory for argument --agent_input_directory")
        sys.exit()
    elif os.path.isdir(known_args.volume_input_directory) == False:
        logging.error(f"Invalid directory for argument --volume_input_directory")
        sys.exit()
    elif os.path.isdir(known_args.agent_archive_directory) == False:
        logging.error(f"Invalid directory for argument --agent_archive_directory")
        sys.exit()
    elif os.path.isdir(known_args.volume_archive_directory) == False:
        logging.error(f"Invalid directory for argument --volume_archive_directory")
        sys.exit()
    # if os.path.isfile(known_args.master_output_file) == False:
    #     print(f"Invalid file for argument --master_output_file, please ensure this argument references a file")
    #     sys.exit()
    else:
        xlsxPath = pathlib.Path(known_args.master_output_file)
        if "xlsx" not in xlsxPath.suffix:
            logging.error(f"Invalid file format for argument --master_output_file, please ensure this argument references a valid xlsx(Excel) file")
            sys.exit()

    _directories = [{"type":"Agent","input":known_args.agent_input_directory, "output":known_args.master_output_file, "archive":known_args.agent_archive_directory},
                    {"type":"Volume","input":known_args.volume_input_directory, "output":known_args.master_output_file, "archive":known_args.volume_archive_directory}]

    for directory in _directories:
        performFileMerge(directory)


if __name__ == '__main__':
  run()
  print("Done")